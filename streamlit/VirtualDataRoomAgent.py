import streamlit as st
import PyPDF2
from docx import Document as DocxDocument
import openpyxl
import io
import os
import re
import pandas as pd
from huggingface_hub import InferenceClient # <-- ADDED
import random # For simulated risk matrix

# --- Helper Functions (from previous version) ---
def extract_text_from_pdf(file_bytes):
    try:
        pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
        text = ""
        for page_num in range(len(pdf_reader.pages)):
            text += pdf_reader.pages[page_num].extract_text()
        if not text.strip():
            return "Text could not be extracted (possibly image-based PDF requiring OCR)."
        return text
    except Exception as e:
        return f"Error extracting PDF text: {str(e)}"

def extract_text_from_docx(file_bytes):
    try:
        doc = DocxDocument(io.BytesIO(file_bytes))
        text = "\n".join([para.text for para in doc.paragraphs])
        return text
    except Exception as e:
        return f"Error extracting DOCX text: {str(e)}"

def extract_text_from_txt(file_bytes):
    try:
        return file_bytes.decode('utf-8', errors='ignore')
    except Exception as e:
        return f"Error extracting TXT text: {str(e)}"

def extract_text_from_xlsx(file_bytes):
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
        text_parts = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            for row in sheet.iter_rows():
                row_text = "\t".join([str(cell.value) if cell.value is not None else "" for cell in row])
                text_parts.append(row_text)
        return "\n".join(text_parts)
    except Exception as e:
        return f"Error extracting XLSX text: {str(e)}"

def get_document_type_from_filename(filename):
    ext = os.path.splitext(filename)[1].lower()
    if ext in ['.pdf', '.doc', '.docx', '.txt']:
        if "agreement" in filename.lower(): return "Agreement"
        if "contract" in filename.lower(): return "Contract"
        if "nda" in filename.lower(): return "Non-disclosure Agreement"
        if "minutes" in filename.lower(): return "Minutes"
        if "financial" in filename.lower() or "statement" in filename.lower(): return "Financial Statement"
        if "articles" in filename.lower(): return "Articles of Incorporation"
        if "bylaws" in filename.lower(): return "Bylaws"
        if "patent" in filename.lower(): return "Patent"
        if "trademark" in filename.lower(): return "Trademark"
        if "lease" in filename.lower(): return "Lease"
        if "deed" in filename.lower(): return "Deed"
        if "employment" in filename.lower(): return "Employment Agreement"
        return "Unknown Document"
    return "Unsupported"

# --- LLM Integration ---
HF_MODEL = "mistralai/Mistral-7B-Instruct-v0.1" # Or another model like "HuggingFaceH4/zephyr-7b-beta"

def get_hf_token():
    # Try to get from Streamlit secrets first (for deployed apps)
    try:
        return st.secrets["HUGGINGFACE_TOKEN"]
    except (FileNotFoundError, KeyError):
        # Fallback to session state if already entered, then to text input
        if "hf_token" in st.session_state and st.session_state.hf_token:
            return st.session_state.hf_token
        return None

def initialize_llm_client(api_token):
    if api_token:
        try:
            return InferenceClient(  provider="hf-inference",
        api_key=st.secrets["hf_token"],)
        except Exception as e:
            st.error(f"Failed to initialize LLM client: {e}")
            return None
    return None

def generate_llm_executive_summary(client, num_docs, key_findings_text):
    if not client:
        return "LLM client not initialized. Please check your Hugging Face API token."
    prompt = f"""
    You are a legal AI assistant helping to draft an executive summary for an M&A due diligence report.
    Based on the following information:
    - Number of documents reviewed (simulated): {num_docs}
    - Key findings and potential risks (simulated extraction): {key_findings_text}

    Please generate a concise executive summary (2-3 paragraphs) highlighting the overall status and critical points for M&A due diligence.
    Focus on a professional tone suitable for legal and business stakeholders.
    Mention that the findings are based on an initial automated review and further human review is essential.
    """
    try:
        response = client.text_generation(prompt, max_new_tokens=300, temperature=0.7)
        return response
    except Exception as e:
        return f"LLM generation error for Executive Summary: {str(e)}"

def generate_llm_detailed_findings_analysis(client, doc_name, clauses, data_points, risks):
    if not client:
        return "LLM client not initialized. Please check your Hugging Face API token."

    clauses_str = "\n".join([f"- {k}: {v}" for k, v in clauses.items()]) if clauses else "None identified."
    data_points_str = "\n".join([f"- {k}: {v}" for k, v in data_points.items()]) if data_points else "None identified."
    risks_str = "\n".join([f"- {k}: {v['status']} - {v['details']}" for k, v in risks.items() if v['status'] == "Potentially Present"]) if risks else "None specifically flagged by initial scan."

    prompt = f"""
    You are a legal AI assistant. For the M&A due diligence document named "{doc_name}", the following information was extracted (simulated):

    Identified Clauses:
    {clauses_str}

    Extracted Data Points:
    {data_points_str}

    Potential Risks Identified:
    {risks_str}

    Please provide a brief analytical paragraph (1-2 paragraphs) summarizing the key implications of these findings for this specific document.
    What should a legal reviewer pay close attention to based on this initial automated extraction?
    Maintain a professional and analytical tone.
    """
    try:
        response = client.text_generation(prompt, max_new_tokens=350, temperature=0.6)
        return response
    except Exception as e:
        return f"LLM generation error for Detailed Findings on {doc_name}: {str(e)}"

def generate_llm_red_flag_analysis(client, red_flags_list):
    if not client:
        return "LLM client not initialized. Please check your Hugging Face API token."
    if not red_flags_list:
        return "No red flags were provided to the LLM for analysis."

    red_flags_str = "\n".join(red_flags_list)
    prompt = f"""
    You are a legal AI risk analyst for M&A due diligence.
    The following potential red flags have been identified across various documents (based on simulated automated extraction):
    {red_flags_str}

    Please provide a short analysis (2-3 paragraphs) of these red flags.
    Discuss the potential implications these types of issues might have in an M&A context.
    Emphasize the need for thorough investigation of each flagged item.
    """
    try:
        response = client.text_generation(prompt, max_new_tokens=300, temperature=0.7)
        return response
    except Exception as e:
        return f"LLM generation error for Red Flag Analysis: {str(e)}"

# --- Main App ---
st.set_page_config(layout="wide", page_title="M&A Due Diligence Assistant")

st.title("⚖️ M&A Due Diligence - Legal Document Analysis System")
st.markdown("""
*As a Legal Technology Specialist, this system assists legal professionals in efficiently reviewing and analyzing large volumes of documents for M&A due diligence.*
**Note:** This is a simulation. True AI capabilities require specialized platforms and trained models. LLM integration provides an example of AI-assisted reporting.
""")

# --- Sidebar for HF Token ---
st.sidebar.header("LLM Configuration")
st.session_state.hf_token = st.sidebar.text_input("Hugging Face API Token", type="password", value=get_hf_token() or "", help="Your Hugging Face API token is required for LLM-generated report sections.")
st.sidebar.caption(f"Using model: `{HF_MODEL}`")

if st.session_state.hf_token:
    if 'llm_client' not in st.session_state or st.session_state.llm_client is None:
        st.session_state.llm_client = initialize_llm_client(st.session_state.hf_token)
        if st.session_state.llm_client:
            st.sidebar.success("LLM Client Initialized!")
        else:
            st.sidebar.error("Failed to initialize LLM Client. Check token or console.")
else:
    st.sidebar.warning("Enter your Hugging Face API token to enable LLM features.")
    st.session_state.llm_client = None


# --- File Uploader ---
uploaded_files = st.file_uploader(
    "Upload relevant legal documents (PDF, DOCX, TXT, XLSX)",
    type=["pdf", "docx", "txt", "xlsx"],
    accept_multiple_files=True
)

if 'documents_data' not in st.session_state:
    st.session_state.documents_data = []
if 'processed_texts' not in st.session_state:
    st.session_state.processed_texts = {}
if 'analysis_results' not in st.session_state: # Stores simulated extraction for each doc
    st.session_state.analysis_results = {}
if 'llm_generated_summary' not in st.session_state:
    st.session_state.llm_generated_summary = ""
if 'llm_generated_detailed_analysis' not in st.session_state:
    st.session_state.llm_generated_detailed_analysis = {} # dict per doc
if 'llm_generated_red_flag_analysis' not in st.session_state:
    st.session_state.llm_generated_red_flag_analysis = ""


if uploaded_files:
    new_files_processed = False
    current_filenames = [doc['name'] for doc in st.session_state.documents_data]
    for uploaded_file in uploaded_files:
        if uploaded_file.name not in current_filenames:
            file_bytes = uploaded_file.getvalue()
            text_content = "N/A"
            file_type = os.path.splitext(uploaded_file.name)[1].lower()

            if file_type == ".pdf":
                text_content = extract_text_from_pdf(file_bytes)
            elif file_type == ".docx":
                text_content = extract_text_from_docx(file_bytes)
            elif file_type == ".txt":
                text_content = extract_text_from_txt(file_bytes)
            elif file_type == ".xlsx":
                text_content = extract_text_from_xlsx(file_bytes)

            doc_info = {
                "name": uploaded_file.name,
                "type": uploaded_file.type,
                "size": uploaded_file.size,
                "doc_type_guess": get_document_type_from_filename(uploaded_file.name),
            }
            st.session_state.documents_data.append(doc_info)
            st.session_state.processed_texts[uploaded_file.name] = text_content
            # Initialize analysis results for new doc
            st.session_state.analysis_results[uploaded_file.name] = {
                'clauses': {}, 'data_points': {}, 'risks': {}
            }
            new_files_processed = True

    if new_files_processed:
        st.success(f"{len(uploaded_files) - len(current_filenames)} new file(s) processed for ingestion.")


if not st.session_state.documents_data:
    st.info("Please upload documents to begin the due diligence process.")
else:
    tab1, tab2, tab3 = st.tabs(["Phase 1: Ingestion & Organization", "Phase 2: Analysis & Extraction", "Phase 3: Collaboration & Reporting"])

    with tab1:
        st.header("Phase 1: Document Ingestion and Organization")
        st.subheader("Task 1: Document Acquisition and Loading")

        if st.session_state.documents_data:
            st.write(f"**{len(st.session_state.documents_data)} document(s) acquired and loaded:**")

            filenames = [doc['name'] for doc in st.session_state.documents_data]
            duplicates = {filename: filenames.count(filename) for filename in filenames if filenames.count(filename) > 1}
            if duplicates:
                st.warning(f"**Duplicate Detection:** Potential duplicates found (by filename): {', '.join(duplicates.keys())}. Real system would use content hashing.")
            else:
                st.success("**Duplicate Detection:** No obvious duplicates found by filename.")

            st.info("System simulates ingestion into a Document Intelligence Platform. OCR would be applied to scanned PDFs and images.")

            cols = st.columns(3)
            cols[0].subheader("Metadata Tagging (Simulated)")
            cols[1].subheader("Folder Structures (Conceptual)")
            cols[2].subheader("Naming Conventions (Conceptual)")

            with cols[0]:
                st.markdown("*Relevant tags would be applied using a consistent taxonomy.*")
                for doc in st.session_state.documents_data:
                    with st.expander(f"Tags for: {doc['name']}"):
                        st.write(f"**Document Type (guessed):** {doc['doc_type_guess']}")
                        st.write(f"**Date (example):** Execution Date: YYYY-MM-DD (Extracted by AI)")
                        # ... (other tags as before)
            with cols[1]:
                st.markdown("""
                *Logical folder structures categorize documents:*
                - By Document Type: `Contracts/`, `Financials/`, `IP/`
                """) # ... (other examples)
            with cols[2]:
                st.markdown("""
                *Consistent file naming conventions ensure clarity:*
                - `[Document Type] - [Party A] vs [Party B] - [Date].pdf`
                """) # ... (other examples)
        else:
            st.write("No documents uploaded yet.")


    with tab2:
        st.header("Phase 2: Document Analysis and Information Extraction")

        if not st.session_state.documents_data:
            st.write("Upload documents in Phase 1 to enable analysis.")
        else:
            st.subheader("Task 2: Key Document Identification")
            search_term = st.text_input("Enter keywords for document search (e.g., 'change of control AND consent')", key="keyword_search_phase2")
            # ... (search logic from previous version) ...

            st.subheader("Task 3: Critical Information Extraction (Simulated AI)")
            st.write("Select a document to see/perform simulated AI extraction:")

            selected_doc_name_for_extraction = st.selectbox(
                "Document for Simulated Extraction",
                options=[doc['name'] for doc in st.session_state.documents_data],
                key="doc_extraction_select_phase2"
            )

            if selected_doc_name_for_extraction:
                doc_text = st.session_state.processed_texts.get(selected_doc_name_for_extraction, "")
                # Ensure analysis_results for this doc exists
                if selected_doc_name_for_extraction not in st.session_state.analysis_results:
                    st.session_state.analysis_results[selected_doc_name_for_extraction] = {
                        'clauses': {}, 'data_points': {}, 'risks': {}
                    }
                
                current_analysis = st.session_state.analysis_results[selected_doc_name_for_extraction]

                with st.expander(f"Simulated AI Extraction for: {selected_doc_name_for_extraction}", expanded=True):
                    st.markdown("**Clause Identification (AI would find these):**")
                    clauses_to_simulate = { # same as before }
                        "Termination Clause": ["terminate", "termination", "notice period", "cure period"],
                        "Change of Control": ["change of control", "change in ownership", "acquisition"],
                        "Indemnification": ["indemnify", "indemnification", "hold harmless"],
                        "MAE Clause": ["material adverse effect", "mae"],
                        "Governing Law": ["governing law", "jurisdiction"],
                        "Confidentiality": ["confidentiality", "confidential information"],
                        "Non-Compete": ["non-compete", "non-competition", "restrictive covenant"],
                        "Force Majeure": ["force majeure", "act of god"]
                    }
                    extracted_clauses = current_analysis.get('clauses', {})
                    for clause_name, keywords in clauses_to_simulate.items():
                        # Only re-simulate if not already done or to show placeholder
                        if clause_name not in extracted_clauses or not extracted_clauses[clause_name]:
                            found_snippet = "[AI: Clause Text Not Found in Simple Search]"
                            if doc_text and doc_text != "Text could not be extracted (possibly image-based PDF requiring OCR).":
                                for kw in keywords:
                                    match = re.search(f"(.{{0,50}}{re.escape(kw)}.{0,100})", doc_text, re.IGNORECASE | re.DOTALL)
                                    if match:
                                        found_snippet = f"...{match.group(1).strip()}..."
                                        break
                            extracted_clauses[clause_name] = found_snippet
                        st.write(f"- **{clause_name}:** `{extracted_clauses[clause_name]}`")
                    current_analysis['clauses'] = extracted_clauses

                    st.markdown("**Data Point Extraction (AI would extract these):**")
                    data_points_to_simulate = { # same as before }
                        "Expiration Date": ["expires on", "expiration date", "term ends"],
                        "Payment Terms": ["payment of", "amount of $", "fee of"],
                        "Renewal Options": ["option to renew", "renewal term"],
                        "Financial Thresholds": ["revenue target of", "minimum purchase of"],
                        "Parties": ["between .* and .*", "Agreement is made .* by"]
                    }
                    extracted_data_points = current_analysis.get('data_points', {})
                    for dp_name, patterns in data_points_to_simulate.items():
                        if dp_name not in extracted_data_points or not extracted_data_points[dp_name]:
                            found_value = "[AI: Data Point Not Found]"
                            if doc_text and doc_text != "Text could not be extracted (possibly image-based PDF requiring OCR).":
                                for pattern in patterns:
                                    match = re.search(pattern, doc_text, re.IGNORECASE)
                                    if match:
                                        found_value = match.group(0)
                                        break
                            extracted_data_points[dp_name] = found_value
                        st.write(f"- **{dp_name}:** `{extracted_data_points[dp_name]}`")
                    current_analysis['data_points'] = extracted_data_points

                    st.markdown("**Risk Factor Identification (AI would flag these):**")
                    risks_to_simulate = { # same as before }
                        "Pending Litigation": ["lawsuit", "litigation", "claim against"],
                        "Regulatory Compliance Violations": ["violation of regulation", "non-compliance", "fine imposed"],
                        "Contractual Breaches": ["breach of contract", "default under this agreement"],
                        "Environmental Liabilities": ["environmental contamination", "remediation cost"],
                        "Employee Liabilities": ["wrongful termination", "discrimination claim"],
                        "Data Privacy Breaches": ["data breach", "privacy violation"]
                    }
                    identified_risks = current_analysis.get('risks', {})
                    for risk_name, keywords in risks_to_simulate.items():
                        if risk_name not in identified_risks or not identified_risks[risk_name]['status']:
                            is_present = False
                            found_details = "[AI: No Specific Details Found]"
                            if doc_text and doc_text != "Text could not be extracted (possibly image-based PDF requiring OCR).":
                                for kw in keywords:
                                    match = re.search(f"(.{{0,50}}{re.escape(kw)}.{0,100})", doc_text, re.IGNORECASE | re.DOTALL)
                                    if match:
                                        is_present = True
                                        found_details = f"...{match.group(1).strip()}..."
                                        break
                            status = "Potentially Present" if is_present else "Not Detected (by simple search)"
                            identified_risks[risk_name] = {"status": status, "details": found_details if is_present else ""}
                        st.write(f"- **{risk_name}:** {identified_risks[risk_name]['status']} `{identified_risks[risk_name]['details'] if identified_risks[risk_name]['status'] == 'Potentially Present' else ''}`")
                    current_analysis['risks'] = identified_risks
                    
                    st.session_state.analysis_results[selected_doc_name_for_extraction] = current_analysis


    with tab3:
        st.header("Phase 3: Collaboration and Reporting")
        if not st.session_state.documents_data:
            st.write("Process documents in Phase 1 & 2 to enable reporting.")
        else:
            st.subheader("Task 4: Data Integration and Collaboration")
            # ... (conceptual description as before) ...
            st.markdown("""
            *Extracted information would be integrated with a legal business management platform (e.g., Thomson Reuters HighQ).*
            - **Database Population:** Data transferred to structured databases (e.g., HighQ iSheets).
            - **Workflow Automation:** Automated review assignments, progress tracking, issue escalation.
            """)

            st.subheader("Data Visualization (Simulated Examples)")
            # ... (visualization code as before, e.g., doc types chart, risk matrix) ...
            if st.session_state.documents_data:
                doc_types = [doc['doc_type_guess'] for doc in st.session_state.documents_data]
                if doc_types:
                    df_doc_types = pd.DataFrame(doc_types, columns=['Document Type']).value_counts().reset_index(name='Count')
                    st.bar_chart(df_doc_types.set_index('Document Type'))

                st.write("**Simulated Risk Matrix / Heatmap (Conceptual)**")
                risk_data_for_matrix = []
                for doc_name_rm, analysis_rm in st.session_state.analysis_results.items():
                    if 'risks' in analysis_rm:
                        for risk_type, risk_info in analysis_rm['risks'].items():
                            if risk_info['status'] == "Potentially Present":
                                likelihood = random.choice(["Low", "Medium", "High"])
                                impact = random.choice(["Low", "Medium", "High"])
                                risk_data_for_matrix.append({
                                    "Document": doc_name_rm, "Risk Type": risk_type,
                                    "Likelihood": likelihood, "Impact": impact,
                                    "Details (Simulated)": risk_info['details']
                                })
                if risk_data_for_matrix:
                    df_risk_matrix = pd.DataFrame(risk_data_for_matrix)
                    st.dataframe(df_risk_matrix)
                else:
                    st.info("No specific risks flagged by simple search to display in matrix.")


            st.subheader("Task 5: Reporting and Risk Surfacing (LLM-Assisted)")
            st.markdown("""
            *Generate comprehensive reports summarizing findings. Use the buttons below to leverage an LLM for specific sections based on the simulated data.*
            """)

            # --- LLM Report Generation Sections ---
            # 1. Executive Summary
            if st.button("Generate Executive Summary with LLM", disabled=not st.session_state.llm_client):
                if st.session_state.llm_client:
                    with st.spinner("LLM is generating Executive Summary..."):
                        # Collect key findings for the prompt
                        num_docs = len(st.session_state.documents_data)
                        
                        all_risks_summary = []
                        for doc_name_es, analysis_es in st.session_state.analysis_results.items():
                            if 'risks' in analysis_es:
                                for risk_type, risk_info in analysis_es['risks'].items():
                                    if risk_info['status'] == "Potentially Present":
                                        all_risks_summary.append(f"- {risk_type} in {doc_name_es}: {risk_info['details']}")
                        key_findings_text = "\n".join(all_risks_summary) if all_risks_summary else "No significant risks highlighted by the initial automated scan."

                        st.session_state.llm_generated_summary = generate_llm_executive_summary(
                            st.session_state.llm_client, num_docs, key_findings_text
                        )
                else:
                    st.warning("LLM client not initialized. Please enter your Hugging Face token in the sidebar.")

            if st.session_state.llm_generated_summary:
                st.markdown("#### LLM-Generated Executive Summary:")
                st.markdown(st.session_state.llm_generated_summary)
                st.markdown("---")

            # 2. Detailed Findings Analysis for a selected document
            st.markdown("#### LLM-Generated Detailed Findings Analysis (Per Document)")
            doc_for_llm_detail = st.selectbox(
                "Select Document for LLM Detailed Analysis",
                options=[doc['name'] for doc in st.session_state.documents_data if doc['name'] in st.session_state.analysis_results],
                key="llm_detail_doc_select"
            )

            if st.button(f"Generate Detailed Analysis for {doc_for_llm_detail} with LLM", disabled=not st.session_state.llm_client or not doc_for_llm_detail):
                if st.session_state.llm_client and doc_for_llm_detail:
                    if doc_for_llm_detail in st.session_state.analysis_results:
                        with st.spinner(f"LLM is analyzing {doc_for_llm_detail}..."):
                            doc_analysis_data = st.session_state.analysis_results[doc_for_llm_detail]
                            st.session_state.llm_generated_detailed_analysis[doc_for_llm_detail] = generate_llm_detailed_findings_analysis(
                                st.session_state.llm_client,
                                doc_for_llm_detail,
                                doc_analysis_data.get('clauses', {}),
                                doc_analysis_data.get('data_points', {}),
                                doc_analysis_data.get('risks', {})
                            )
                    else:
                        st.warning(f"No analysis data found for {doc_for_llm_detail}. Please run extraction in Phase 2.")
                else:
                    st.warning("LLM client not initialized or no document selected.")
            
            if doc_for_llm_detail and doc_for_llm_detail in st.session_state.llm_generated_detailed_analysis:
                st.markdown(f"**Analysis for {doc_for_llm_detail}:**")
                st.markdown(st.session_state.llm_generated_detailed_analysis[doc_for_llm_detail])
                st.markdown("---")

            # 3. Red Flag Report Analysis
            if st.button("Generate Red Flag Analysis with LLM", disabled=not st.session_state.llm_client):
                if st.session_state.llm_client:
                    with st.spinner("LLM is generating Red Flag Analysis..."):
                        red_flags_list = []
                        for doc_name_rf, analysis_rf in st.session_state.analysis_results.items():
                            if 'risks' in analysis_rf:
                                for risk_type, risk_info in analysis_rf['risks'].items():
                                    if risk_info['status'] == "Potentially Present":
                                        red_flags_list.append(f"Document '{doc_name_rf}': Potential {risk_type} - Details: {risk_info['details']}")
                        
                        if red_flags_list:
                            st.session_state.llm_generated_red_flag_analysis = generate_llm_red_flag_analysis(
                                st.session_state.llm_client, red_flags_list
                            )
                        else:
                            st.session_state.llm_generated_red_flag_analysis = "No red flags were identified by the simulated scan to provide to the LLM for analysis."
                else:
                    st.warning("LLM client not initialized.")
            
            if st.session_state.llm_generated_red_flag_analysis:
                st.markdown("#### LLM-Generated Red Flag Analysis:")
                st.markdown(st.session_state.llm_generated_red_flag_analysis)
                st.markdown("---")


            # --- Downloadable Report ---
            st.subheader("Download Full Report (Simulated + LLM)")
            report_content = []
            report_content.append("# M&A Due Diligence - LLM-Assisted Report\n\n")

            if st.session_state.llm_generated_summary:
                report_content.append("## LLM-Generated Executive Summary\n")
                report_content.append(st.session_state.llm_generated_summary + "\n\n")
            else:
                report_content.append("## Executive Summary (Simulated - LLM not run or failed)\n")
                report_content.append(f"Review of {len(st.session_state.documents_data)} documents initiated. (Placeholder)\n\n")

            report_content.append("## Detailed Findings (Simulated Extraction & LLM Analysis)\n")
            if st.session_state.analysis_results:
                for doc_name_rep, results_rep in st.session_state.analysis_results.items():
                    report_content.append(f"### Document: {doc_name_rep}\n")
                    if 'clauses' in results_rep and results_rep['clauses']:
                        report_content.append("#### Simulated Key Clauses Identified:\n")
                        for clause, text in results_rep['clauses'].items():
                            report_content.append(f"- **{clause}:** {text}\n")
                    if 'data_points' in results_rep and results_rep['data_points']:
                        report_content.append("\n#### Simulated Key Data Points Extracted:\n")
                        for dp, val in results_rep['data_points'].items():
                            report_content.append(f"- **{dp}:** {val}\n")
                    
                    if doc_name_rep in st.session_state.llm_generated_detailed_analysis:
                        report_content.append("\n#### LLM-Generated Analysis for this Document:\n")
                        report_content.append(st.session_state.llm_generated_detailed_analysis[doc_name_rep] + "\n")
                    report_content.append("---\n")
            else:
                report_content.append("No detailed analysis performed or available for the report yet.\n\n")

            report_content.append("## Red Flag Report\n")
            if st.session_state.llm_generated_red_flag_analysis:
                report_content.append("### LLM-Generated Red Flag Analysis:\n")
                report_content.append(st.session_state.llm_generated_red_flag_analysis + "\n\n")
            else:
                report_content.append("### Simulated Red Flags (LLM for analysis not run or failed):\n")
                high_risk_items = []
                for doc_name_hr, analysis_hr in st.session_state.analysis_results.items():
                    if 'risks' in analysis_hr:
                        for risk_type, risk_info in analysis_hr['risks'].items():
                            if risk_info['status'] == "Potentially Present":
                                high_risk_items.append(f"- **{risk_type}** in *{doc_name_hr}*: {risk_info['details']}\n")
                if high_risk_items:
                    report_content.extend(high_risk_items)
                else:
                    report_content.append("No specific red flags identified by simple search in analyzed documents.\n\n")

            report_content.append("## Review Status Report (Simulated)\n")
            report_content.append(f"- **Total Documents Uploaded:** {len(st.session_state.documents_data)}\n")
            report_content.append(f"- **Documents with Simulated AI Analysis:** {len([d for d in st.session_state.analysis_results if st.session_state.analysis_results[d].get('clauses')])}\n")
            # ... (other status metrics) ...

            final_report_str = "".join(report_content)

            st.download_button(
                label="Download Full Report (Markdown)",
                data=final_report_str,
                file_name="llm_assisted_due_diligence_report.md",
                mime="text/markdown"
            )
            with st.expander("View Full Report Content"):
                st.markdown(final_report_str)

# --- Footer ---
st.sidebar.info("LegalTech Specialist AI Agent for M&A Due Diligence")
# ... (rest of the sidebar as before) ...